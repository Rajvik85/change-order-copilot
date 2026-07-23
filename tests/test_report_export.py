"""Excel and self-contained HTML export tests."""

from __future__ import annotations

import json
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from co_copilot import __version__
from co_copilot.models import LLMResult
from co_copilot.report_export import (
    compliance_json_bytes,
    excel_register_bytes,
    extracted_json_bytes,
    html_review_pack,
)


def test_excel_contains_three_formatted_sheets(demo_result) -> None:
    workbook = load_workbook(BytesIO(excel_register_bytes(demo_result, __version__)))
    assert workbook.sheetnames == [
        "CO Register",
        "Portfolio Summary",
        "Compliance Scorecard",
    ]
    register = workbook["CO Register"]
    assert register.freeze_panes == "A2"
    assert register.auto_filter.ref
    assert register["A1"].font.bold


def test_html_pack_is_self_contained_and_ai_labeled(demo_result) -> None:
    memo = LLMResult(status="ok", content="Grounded example", provider="mock")
    pack = html_review_pack(
        demo_result,
        __version__,
        date(2026, 7, 23),
        "Hold commercial meeting.",
        memos={"CO-001": memo},
    )
    assert "<!doctype html>" in pack.lower()
    assert "AI-assisted draft — human review required" in pack
    assert "Hold commercial meeting" in pack
    assert "CO-018" in pack
    assert "https://" not in pack
    assert demo_result.config_hash in pack


def test_machine_readable_exports_include_manifest(demo_result) -> None:
    facts = json.loads(extracted_json_bytes(demo_result, __version__))
    compliance = json.loads(compliance_json_bytes(demo_result, __version__))
    assert facts["manifest"]["input_data_hash"] == demo_result.input_hash
    assert len(facts["extractions"]) == 18
    assert len(compliance["compliance"]) == 18
