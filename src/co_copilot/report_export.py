"""Traceable Excel and self-contained HTML review-pack exports."""

from __future__ import annotations

import html
import io
import json
from dataclasses import asdict
from datetime import date

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from co_copilot.models import LLMResult, PipelineResult
from co_copilot.pipeline import result_manifest

_NAVY = "10243E"
_AMBER = "D89414"
_GREEN = "B7E1CD"
_YELLOW = "FFF2CC"
_RED = "F4CCCC"


def _format_sheet(worksheet, freeze: str = "A2") -> None:
    """Apply readable register formatting without relying on Excel macros."""
    worksheet.freeze_panes = freeze
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for column in worksheet.columns:
        values = [str(cell.value or "") for cell in column[:100]]
        width = min(max(max(map(len, values), default=8) + 2, 11), 38)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = width


def excel_register_bytes(result: PipelineResult, version: str) -> bytes:
    """Create the formatted three-sheet commercial register workbook."""
    buffer = io.BytesIO()
    register = result.analytics["register"].copy()
    summary = pd.DataFrame(
        [
            ("Application version", version),
            ("Generated at", result.generated_at),
            ("Configuration hash", result.config_hash),
            ("Input-data hash", result.input_hash),
            ("Provisional", result.provisional),
            ("Documents", len(result.documents)),
            ("Net exposure", result.analytics["total_exposure"]),
            ("Gross positive exposure", result.analytics["gross_positive_exposure"]),
            ("Exposure % of contract sum", result.analytics["exposure_percent"]),
            ("Timely notice rate", result.analytics["compliance_rate"]),
            ("Time-barred gross value", result.analytics["time_barred_value"]),
        ],
        columns=["Metric", "Value"],
    )
    scorecard = pd.DataFrame(
        [
            {
                "CO number": item.co_number,
                "Time-bar verdict": item.time_bar_verdict,
                "Notice elapsed days": item.notice_elapsed_days,
                "Particulars elapsed days": item.particulars_elapsed_days,
                "Compliance score": item.score,
                **{check.label: check.status for check in item.checks},
            }
            for item in result.compliance
        ]
    )
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        register.to_excel(writer, sheet_name="CO Register", index=False)
        summary.to_excel(writer, sheet_name="Portfolio Summary", index=False)
        scorecard.to_excel(writer, sheet_name="Compliance Scorecard", index=False)
        for worksheet in writer.book.worksheets:
            _format_sheet(worksheet)
        register_sheet = writer.book["CO Register"]
        verdict_column = (
            list(register.columns).index("time_bar_verdict") + 1
            if "time_bar_verdict" in register
            else None
        )
        if verdict_column:
            letter = get_column_letter(verdict_column)
            register_sheet.conditional_formatting.add(
                f"{letter}2:{letter}{register_sheet.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=['"COMPLIANT"'],
                    fill=PatternFill("solid", fgColor=_GREEN),
                ),
            )
            register_sheet.conditional_formatting.add(
                f"{letter}2:{letter}{register_sheet.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=['"AT RISK"'],
                    fill=PatternFill("solid", fgColor=_YELLOW),
                ),
            )
            register_sheet.conditional_formatting.add(
                f"{letter}2:{letter}{register_sheet.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=['"TIME-BARRED"'],
                    fill=PatternFill("solid", fgColor=_RED),
                ),
            )
        for sheet_name in ("CO Register", "Compliance Scorecard"):
            worksheet = writer.book[sheet_name]
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return buffer.getvalue()


def _money(value: object) -> str:
    return f"USD {float(value or 0):,.0f}"


def _verdict_class(verdict: str) -> str:
    return {
        "COMPLIANT": "pass",
        "PASS": "pass",
        "AT RISK": "warn",
        "WARN": "warn",
        "TIME-BARRED": "fail",
        "FAIL": "fail",
        "INFO": "info",
    }.get(verdict, "info")


def html_review_pack(
    result: PipelineResult,
    version: str,
    as_of_date: date,
    reviewer_narrative: str = "",
    memos: dict[str, LLMResult] | None = None,
    sections: set[str] | None = None,
) -> str:
    """Build a self-contained HTML review pack with no remote dependencies."""
    memos = memos or {}
    sections = sections or {"summary", "findings", "co_pages", "method"}
    manifest = result_manifest(result, version)
    register = result.analytics["register"]
    cards = ""
    by_result = {item.document_id: item for item in result.compliance}
    extraction_by_id = {item.document_id: item for item in result.extractions}
    if "co_pages" in sections:
        card_parts: list[str] = []
        for row in register.itertuples():
            assessment = by_result[row.document_id]
            extraction = extraction_by_id[row.document_id]
            checks = "".join(
                f"<li class='{_verdict_class(check.status)}'><strong>"
                f"{html.escape(check.label)} — {html.escape(check.verdict)}</strong>"
                f"<br>{html.escape(check.calculation)}"
                f"<small>{html.escape(check.rule)}</small></li>"
                for check in assessment.checks
            )
            memo = memos.get(row.document_id)
            memo_html = ""
            if memo and memo.status == "ok":
                memo_html = (
                    "<div class='ai-label'>AI-assisted draft — human review required</div>"
                    f"<div class='memo'>{html.escape(memo.content).replace(chr(10), '<br>')}</div>"
                )
            clauses = (
                ", ".join(extraction.value("cited_clause_refs", [])) or "Not cited"
            )
            card_parts.append(
                f"""
                <article class="co-card">
                  <div class="co-title"><h2>{html.escape(row.co_number)}</h2>
                    <span class="badge {_verdict_class(assessment.time_bar_verdict)}">
                    {html.escape(assessment.time_bar_verdict)}</span></div>
                  <div class="facts">
                    <div><small>Type</small><strong>{html.escape(str(row.type))}</strong></div>
                    <div><small>Contractor</small><strong>{html.escape(str(row.contractor))}</strong></div>
                    <div><small>Value</small><strong>{_money(row.cost_value)}</strong></div>
                    <div><small>Schedule</small><strong>{row.schedule_days:+d} days</strong></div>
                    <div><small>Status</small><strong>{html.escape(str(row.status).title())}</strong></div>
                    <div><small>Clauses</small><strong>{html.escape(clauses)}</strong></div>
                  </div>
                  <h3>Deterministic compliance checklist</h3><ul>{checks}</ul>
                  {memo_html}
                </article>
                """
            )
        cards = "\n".join(card_parts)

    findings = ""
    if "findings" in sections:
        findings = "".join(
            f"<div class='finding {item.severity}'><strong>{html.escape(item.title)}</strong>"
            f"<p>{html.escape(item.message)}</p><small>{html.escape(item.rationale)}</small></div>"
            for item in result.analytics["findings"]
        )
    summary = ""
    if "summary" in sections:
        summary = f"""
        <section class="summary">
          <div><small>Net exposure</small><strong>{_money(result.analytics['total_exposure'])}</strong></div>
          <div><small>Contract share</small><strong>{result.analytics['exposure_percent']:.1f}%</strong></div>
          <div><small>Timely notices</small><strong>{result.analytics['compliance_rate']:.0%}</strong></div>
          <div><small>Time-barred gross</small><strong>{_money(result.analytics['time_barred_value'])}</strong></div>
        </section>
        """
    methodology = ""
    if "method" in sections:
        methodology = (
            "<section><h2>Method and limitations</h2><p>Facts were extracted by the "
            "offline deterministic NLP core. Compliance verdicts use configured "
            "calendar-day rules and clause-to-change mappings. Claimed schedule days "
            "are shown as exposure indicators and are not assumed additive or critical.</p></section>"
        )
    narrative = (
        f"<section><h2>Reviewer narrative</h2><p>{html.escape(reviewer_narrative)}</p></section>"
        if reviewer_narrative.strip()
        else ""
    )
    manifest_json = html.escape(json.dumps(manifest, indent=2))
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>CO Review Pack · {html.escape(str(as_of_date))}</title>
<style>
body{{font-family:Inter,Segoe UI,Arial,sans-serif;margin:0;background:#f5f7fa;color:#10243e}}
main{{max-width:1080px;margin:auto;padding:36px 24px 64px}} header{{background:#10243e;color:white;padding:34px;border-radius:18px}}
h1{{margin:.2rem 0}} h2,h3{{color:#10243e}} small{{display:block;color:#64758a;margin-top:4px}}
.summary,.facts{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:22px 0}}
.summary div,.facts div{{background:white;border:1px solid #dbe4ed;border-radius:10px;padding:14px}}
.summary strong,.facts strong{{display:block;margin-top:7px}} .co-card{{background:white;border:1px solid #dbe4ed;border-radius:16px;padding:22px;margin:20px 0;page-break-inside:avoid}}
.co-title{{display:flex;align-items:center;justify-content:space-between}} .badge{{padding:6px 10px;border-radius:999px;font-weight:700}}
.pass{{color:#006b54}} .warn{{color:#7a5600}} .fail{{color:#a12f20}} .info{{color:#0072b2}}
.badge.pass{{background:#d7f0e7}} .badge.warn{{background:#fff1c8}} .badge.fail{{background:#f9d9d4}}
li{{margin:12px 0;padding-left:8px}} .finding{{background:white;border-left:6px solid #0072b2;padding:14px;margin:10px 0}}
.finding.critical{{border-color:#a12f20}} .finding.high{{border-color:#d89414}} .ai-label{{background:#fff1c8;border:1px solid #d89414;padding:8px;font-weight:700}}
.memo{{white-space:normal;padding:14px;background:#f7f9fc}} footer{{color:#64758a;font-size:12px;margin-top:30px}}
@media(max-width:700px){{.summary,.facts{{grid-template-columns:1fr 1fr}}}} @media print{{body{{background:white}}}}
</style></head><body><main>
<header><small style="color:#f3bd55">SYNTHETIC PORTFOLIO REVIEW</small>
<h1>Change Order Review Pack</h1><p>As of {html.escape(str(as_of_date))} · {len(register)} change orders</p></header>
{summary}{narrative}<section><h2>Ranked findings</h2>{findings}</section>{cards}{methodology}
<footer>Change Order Copilot v{html.escape(version)} · Generated {html.escape(result.generated_at)}
 · Config {html.escape(result.config_hash)} · Input {html.escape(result.input_hash)}
<details><summary>Traceability manifest</summary><pre>{manifest_json}</pre></details></footer>
</main></body></html>"""


def extracted_json_bytes(result: PipelineResult, version: str) -> bytes:
    """Export canonical facts and traceability metadata as UTF-8 JSON."""
    payload = {
        "manifest": result_manifest(result, version),
        "extractions": [item.to_dict() for item in result.extractions],
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def compliance_json_bytes(result: PipelineResult, version: str) -> bytes:
    """Export the deterministic checklist as UTF-8 JSON."""
    payload = {
        "manifest": result_manifest(result, version),
        "compliance": [asdict(item) for item in result.compliance],
    }
    return json.dumps(payload, indent=2, default=str).encode("utf-8")
